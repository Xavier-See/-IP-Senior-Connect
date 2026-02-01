import paho.mqtt.client as mqtt
import json
import time
import smtplib
import base64
import socket
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from openpyxl import Workbook, load_workbook
import os
import threading
import subprocess

# ============================
# 🔒 EXCEL THREAD LOCK
# ============================
excel_lock = threading.Lock()
sensor_lock = threading.Lock()

# --- CONSOLE COLORS (ANSI) ---
RED = "\033[91m"
GREEN = "\033[92m"
YELLOW = "\033[93m"
ORANGE = "\033[38;5;208m"
CYAN = "\033[96m"
RESET = "\033[0m"

# --- CONFIGURATION ---
BROKER_ADDRESS = "raspberrypi.local"
TOPIC = "senior_connect/sensors/#"

# --- EMAIL SETTINGS ---
SENDER_EMAIL = "cargvr825@gmail.com"
APP_PASSWORD = "mpmr zvpy pgxu cbjj"
RECEIVER_EMAIL = "carereciever825@gmail.com"
WORKFLOW_EMAIL = "carereciever825@gmail.com"

# --- REPORT SETTINGS ---
REPORT_INTERVAL = 60

# ============================
# ✅ FIXED EXCEL/GIT BASE PATH
# ============================
BASE_DIR = "/home/admin/seniorconnect_repo"
EXCEL_FILENAME = "SeniorConnect_MasterLog.xlsx"
EXCEL_PATH = os.path.join(BASE_DIR, EXCEL_FILENAME)

# ============================
# 🛌 BEDROOM LOW-VITALS ALERT SETTINGS
# ============================
BEDROOM_VITALS_HR_LOW = 70          # ✅ CHANGED from 45 -> 49 (as requested)
BEDROOM_VITALS_BR_LOW = 5
BEDROOM_VITALS_ALERT_COOLDOWN = 120
BEDROOM_VITALS_CONFIRM_SECONDS = 30  # ✅ NEW: wait 30s before sending alert

# ============================
# ✅ NEW: BROKER RESOLUTION FIX
# ============================
def resolve_broker_address(addr: str) -> str:
    try:
        return socket.gethostbyname(addr)
    except Exception:
        pass
    if addr.endswith(".local"):
        try_addr = addr.replace(".local", "")
        try:
            return socket.gethostbyname(try_addr)
        except Exception:
            pass
    env_ip = os.getenv("BROKER_IP", "").strip()
    if env_ip:
        return env_ip
    return addr

# ============================
# ✅ NEW: BEDROOM VITALS EMAIL BUILDER
# ============================
def build_bedroom_vitals_email(hr_val, br_val, hr_low, br_low):
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    issues = []
    if hr_low:
        issues.append(f"Low Heart Rate (HR: {hr_val} bpm, threshold < {BEDROOM_VITALS_HR_LOW})")
    if br_low:
        issues.append(f"Low Breathing Rate (BR: {br_val} breaths/min, threshold < {BEDROOM_VITALS_BR_LOW})")
    issue_line = "; ".join(issues) if issues else "Abnormal vital signs detected"
    body = (
        "Dear Caregiver,\n\n"
        "This is an automated alert from the Senior Connect monitoring system.\n"
        "Potential medical emergency indicators were detected in the Bedroom and require immediate attention.\n\n"
        "Alert Details:\n"
        f"- Location: Bedroom\n"
        f"- Time: {now_str}\n"
        f"- Condition: {issue_line}\n"
        f"- Heart Rate (HR): {hr_val if hr_val is not None else 'N/A'} bpm\n"
        f"- Breathing Rate (BR): {br_val if br_val is not None else 'N/A'} breaths/min\n\n"
        "Recommended Actions:\n"
        "1) Check on the senior immediately.\n"
        "2) If unresponsive or symptoms appear serious, call emergency services.\n"
        "3) Continue monitoring for further updates.\n\n"
        "Regards,\n"
        "Senior Connect Alert System"
    )
    return body

# --- GLOBAL STATE TRACKING ---
room_states = {
    "Living Room": {"last_motion": time.time(), "last_log_time": 0, "last_fall_alert": 0},
    "Bedroom": {
        "last_motion": time.time(),
        "last_log_time": 0,
        "last_fall_alert": 0,
        "is_occupied": False,
        "entry_time": time.time(),
        "door_debounce": 0,
        "last_vitals_alert": 0,
        "low_vitals_pending": False,
        "low_vitals_start": 0
    },
    "Bathroom": {
        "is_occupied": False,
        "entry_time": time.time(),
        "last_inside_motion": time.time(),
        "humidity": 0,
        "door_debounce": 0,
        "high_humidity_start": 0,
        "is_high_humidity": False,
        "humidity_alert_sent": False,
        "alert_level": None,
        "critical_sent": False,
        "last_fall_alert": 0
    }
}
entrance_state = {"last_exit_time": 0, "is_away": False}

# ============================
# 📊 EXCEL LOGGING
# ============================
def log_to_excel(sensor_type, location, value, status):
    with excel_lock:
        os.makedirs(BASE_DIR, exist_ok=True)
        now = datetime.now()
        row = [now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"), now.strftime("%H:00"), location, value, status]
        headers = ["Date", "Timestamp", "Hour", "Location", "Value", "Status"]

        sheet_name = sensor_type
        if "Access" in str(sensor_type) or "Entrance" in str(sensor_type):
            sheet_name = "Proximity"
        if sheet_name == "System":
            sheet_name = None

        if not os.path.exists(EXCEL_PATH):
            wb = Workbook()
            wb.active.title = "ALERTS"
            wb["ALERTS"].append(headers)
            for s in ["PIR", "Humidity", "Temperature", "Proximity", "mmWave", "Camera", "mmWave(BR)", "mmWave(HR)", "mmWave(InBed)"]:
                if s not in wb.sheetnames:
                    ws = wb.create_sheet(s)
                    ws.append(headers)
            wb.save(EXCEL_PATH)

        try:
            wb = load_workbook(EXCEL_PATH)
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    ws_new = wb.create_sheet(sheet_name)
                    ws_new.append(headers)
                wb[sheet_name].append(row)

            if status and any(x in str(status).upper() for x in ["ALERT", "WARNING", "CRITICAL", "MINIMAL", "MODERATE"]):
                if "ALERTS" not in wb.sheetnames:
                    ws_alerts = wb.create_sheet("ALERTS")
                    ws_alerts.append(headers)
                wb["ALERTS"].append(row)

            wb.save(EXCEL_PATH)
            print(f"✅ [LOGGED] {sheet_name if sheet_name else 'ALERTS'} -> {location} ({value})")
        except Exception as e:
            print(f"❌ [ERROR] Excel Save Failed: {e}")

# ============================
# 📧 SEND EXCEL REPORT
# ============================
def send_excel_report():
    if not os.path.exists(EXCEL_PATH):
        print("⚠️ No Excel file to send yet.")
        return
    try:
        msg = MIMEMultipart()
        msg["From"] = SENDER_EMAIL
        msg["To"] = WORKFLOW_EMAIL
        msg["Subject"] = f"📊 Senior Connect Data Log: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        msg.attach(MIMEText("Attached is the latest sensor data log for the Workflow Team.", "plain"))

        with open(EXCEL_PATH, "rb") as f:
            part = MIMEApplication(f.read(), Name=EXCEL_FILENAME)
            part['Content-Disposition'] = f'attachment; filename="{EXCEL_FILENAME}"'
            msg.attach(part)

        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(SENDER_EMAIL, APP_PASSWORD)
        server.sendmail(SENDER_EMAIL, WORKFLOW_EMAIL, msg.as_string())
        server.quit()
        print(f"📤 [REPORT SENT] Excel file sent to {WORKFLOW_EMAIL}")
    except Exception as e:
        print(f"❌ [REPORT ERROR] Could not send excel: {e}")

# ============================
# 🔁 PUSH TO GITHUB (MAIN BRANCH) - QUIET
# ============================
def push_excel_to_github():
    if not os.path.exists(EXCEL_PATH):
        return
    try:
        quiet = {"cwd": BASE_DIR, "check": True, "stdout": subprocess.DEVNULL, "stderr": subprocess.DEVNULL}
        subprocess.run(["git", "add", EXCEL_FILENAME], **quiet)
        subprocess.run(["git", "commit", "-m", "Auto update Excel log"], cwd=BASE_DIR, check=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        subprocess.run(["git", "pull", "--no-rebase", "origin", "main"], **quiet)
        subprocess.run(["git", "push", "origin", "main"], **quiet)
        print("✅ Excel pushed to GitHub (main branch)")
    except subprocess.CalledProcessError as e:
        print("❌ Git push failed:", e)

# ============================
# EMAIL ALERT
# ============================
def send_email_alert(subject, body):
    try:
        msg = MIMEMultipart()
        msg["From"] = SENDER_EMAIL
        msg["To"] = RECEIVER_EMAIL
        msg["Subject"] = f"⚠️ SENIOR CONNECT: {subject}"
        msg.attach(MIMEText(body, "plain"))

        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(SENDER_EMAIL, APP_PASSWORD)
        server.sendmail(SENDER_EMAIL, RECEIVER_EMAIL, msg.as_string())
        server.quit()
        print(f"{GREEN}📧 [EMAIL SENT] {subject}{RESET}")
    except Exception as e:
        print(f"❌ [EMAIL ERROR] {e}")

# ============================
# BATHROOM LOGIC
# ============================
def check_bathroom_logic():
    state = room_states["Bathroom"]
    now = time.time()
    if not state["is_occupied"]:
        state["alert_level"] = None
        state["critical_sent"] = False
        state["humidity_alert_sent"] = False
        return

    time_inside = now - state["entry_time"]
    time_no_motion = now - state["last_inside_motion"]

    HUMIDITY_THRESHOLD = 90.0
    HUMIDITY_DURATION_LIMIT = 20

    if time_no_motion < 2:
        state["high_humidity_start"] = now

    if state["humidity"] > HUMIDITY_THRESHOLD:
        if not state["is_high_humidity"]:
            state["is_high_humidity"] = True
            state["high_humidity_start"] = now
            print(f"{CYAN}💧 High Humidity Detected ({state['humidity']}%) - Timer Started{RESET}")
        else:
            duration = now - state["high_humidity_start"]
            if duration > HUMIDITY_DURATION_LIMIT and not state["humidity_alert_sent"] and time_no_motion > 5:
                msg = f"Bathroom Humidity > {HUMIDITY_THRESHOLD}% for {int(duration)}s with NO MOTION. Alerting."
                print(f"{CYAN}🌊 {msg} -> Sending Alert{RESET}")
                send_email_alert("High Humidity Alert", msg)
                log_to_excel("System", "Bathroom", "WARNING", msg)
                state["humidity_alert_sent"] = True
    else:
        if state["is_high_humidity"]:
            print(f"{GREEN}📉 Humidity Normalized ({state['humidity']}%) - Timer Reset{RESET}")
            state["is_high_humidity"] = False
            state["high_humidity_start"] = 0
            state["humidity_alert_sent"] = False

    if time_no_motion >= 20 and state["alert_level"] is None:
        msg = "MINIMAL ALERT: Bathroom occupied, no motion > 20s."
        print(f"{YELLOW}⚠️ {msg}{RESET}")
        send_email_alert("Bathroom Minimal Alert", msg)
        log_to_excel("System", "Bathroom", "MINIMAL", msg)
        state["alert_level"] = "minimal"
    elif time_no_motion >= 40 and state["alert_level"] == "minimal":
        msg = "MODERATE ALERT: Bathroom occupied, no motion > 40s."
        print(f"{ORANGE}🚨 {msg}{RESET}")
        send_email_alert("Bathroom Moderate Alert", msg)
        log_to_excel("System", "Bathroom", "MODERATE", msg)
        state["alert_level"] = "moderate"
    elif time_no_motion >= 60 and not state["critical_sent"]:
        msg = "CRITICAL ALERT: Bathroom occupied, no motion > 60s."
        print(f"{RED}🆘 {msg}{RESET}")
        send_email_alert("Bathroom CRITICAL Alert", msg)
        log_to_excel("System", "Bathroom", "CRITICAL", msg)
        state["alert_level"] = "critical"
        state["critical_sent"] = True

# ================= THREAD LOOPS =================
def bathroom_monitor_loop():
    while True:
        check_bathroom_logic()
        time.sleep(1)

def report_monitor_loop():
    print(f"⏳ Reporting Scheduler Started (Interval: {REPORT_INTERVAL}s)")
    while True:
        time.sleep(REPORT_INTERVAL)
        print("📤 Preparing to send scheduled Excel report...")
        with excel_lock:
            send_excel_report()
            push_excel_to_github()

# ================= MQTT CALLBACKS =================
def on_connect(client, userdata, flags, rc, properties=None):
    if rc == 0:
        print("🟢 MQTT Connected")
        client.subscribe(TOPIC)

def send_image_email_threaded(image_bytes, filename="Entrance.jpg"):
    """Send image via email in a separate thread to avoid blocking main loop."""
    def worker():
        try:
            msg = MIMEMultipart()
            msg["From"] = SENDER_EMAIL
            msg["To"] = RECEIVER_EMAIL
            msg["Subject"] = f"📸 Entrance Image: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            msg.attach(MIMEText("Attached is the latest Entrance image.", "plain"))

            part = MIMEApplication(image_bytes, Name=filename)
            part['Content-Disposition'] = f'attachment; filename="{filename}"'
            msg.attach(part)

            server = smtplib.SMTP("smtp.gmail.com", 587)
            server.starttls()
            server.login(SENDER_EMAIL, APP_PASSWORD)
            server.sendmail(SENDER_EMAIL, RECEIVER_EMAIL, msg.as_string())
            server.quit()
            print(f"📤 Email sent: {filename}")
        except Exception as e:
            print(f"❌ Email error: {e}")
    threading.Thread(target=worker, daemon=True).start()

def on_message(client, userdata, msg):
    try:
        if msg.retain == 1:
            return
        payload = json.loads(msg.payload.decode())
        s_type = payload.get("type")
        loc = payload.get("location")
        val = payload.get("value")
        now = time.time()

        # --- BATHROOM MOTION ---
        if s_type == "PIR" and loc == "Bathroom":
            if room_states["Bathroom"]["is_occupied"]:
                if "Motion" in str(val) and "No" not in str(val):
                    room_states["Bathroom"]["last_inside_motion"] = now
                    if not room_states["Bathroom"]["critical_sent"]:
                        room_states["Bathroom"]["alert_level"] = None
                    print("🚶 Bathroom Motion - Timer Reset")
            log_to_excel(s_type, loc, val, payload.get("status"))

        # --- BATHROOM DOOR ---
        elif s_type == "Proximity" and loc == "Bathroom Door":
            with sensor_lock:
                bt = room_states["Bathroom"]
                DEBOUNCE_TIME = 2
                last_debounce = bt.get("door_debounce", 0)
                log_value = None

                if str(val).upper() in ["ENTER", "DETECTED"]:
                    if not bt["is_occupied"] and (now - last_debounce > DEBOUNCE_TIME):
                        bt["is_occupied"] = True
                        bt["entry_time"] = now
                        bt["last_inside_motion"] = now
                        bt["alert_level"] = None
                        print(f"{GREEN}🚪 Bathroom ENTER{RESET}")
                        bt["door_debounce"] = now
                        log_value = "ENTER"
                elif str(val).upper() in ["EXIT", "CLEAR"]:
                    if bt["is_occupied"] and (now - last_debounce > DEBOUNCE_TIME):
                        if now - bt["entry_time"] > 2:
                            bt["is_occupied"] = False
                            bt["alert_level"] = None
                            bt["critical_sent"] = False
                            bt["humidity_alert_sent"] = False
                            print(f"{RED}🚪 Bathroom EXIT{RESET}")
                            bt["door_debounce"] = now
                            log_value = "EXIT"

                if log_value:
                    log_to_excel("Proximity", "Bathroom Door", log_value, f"User {log_value}ED Bathroom")
                room_states["Bathroom"] = bt

        # --- BEDROOM DOOR ---
        elif s_type == "Proximity" and loc in ["Bedroom Door", "Bedroom"]:
            with sensor_lock:
                bd = room_states["Bedroom"]
                DEBOUNCE_TIME = 2
                last_debounce = bd.get("door_debounce", 0)
                log_value = None

                if str(val).upper() in ["ENTER", "DETECTED", "BLOCKED"]:
                    if not bd.get("is_occupied", False) and (now - last_debounce > DEBOUNCE_TIME):
                        if now - bd["entry_time"] > 2:
                            bd["is_occupied"] = True
                            bd["entry_time"] = now
                            bd["door_debounce"] = now
                            log_value = "ENTER"
                            print(f"{GREEN}🚪 Bedroom ENTER{RESET}")
                elif str(val).upper() in ["EXIT", "CLEAR"]:
                    if bd.get("is_occupied", False) and (now - last_debounce > DEBOUNCE_TIME):
                        if now - bd["entry_time"] > 2:
                            bd["is_occupied"] = False
                            bd["door_debounce"] = now
                            log_value = "EXIT"
                            print(f"{RED}🚪 Bedroom EXIT{RESET}")
                            bd["low_vitals_pending"] = False
                            bd["low_vitals_start"] = 0

                if log_value:
                    log_to_excel("Proximity", "Bedroom Door", log_value, f"User {log_value}ED Bedroom")
                room_states["Bedroom"] = bd

        # --- CAMERA: LIVING ROOM MAIN DOOR ---
        elif s_type == "Camera" and loc == "Living Room Main Door":
            try:
                image_data = base64.b64decode(payload["image"])
                log_to_excel("Camera", loc, "Image Captured", "Active")
                send_image_email_threaded(
                    image_data,
                    filename=f"LivingRoom_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                )
            except Exception as e:
                print(f"❌ Error processing camera image: {e}")

        # --- LIVING ROOM DOOR ---
        elif s_type == "Proximity" and loc == "Living Room Main Door":
            with sensor_lock:
                bt = room_states.get(loc, {})
                DEBOUNCE_TIME = 2
                last_debounce = bt.get("door_debounce", 0)
                log_value = None

                if str(val).upper() in ["ENTER", "DETECTED", "BLOCKED"]:
                    if not bt.get("is_occupied", False) and (now - last_debounce > DEBOUNCE_TIME):
                        if now - bt.get("entry_time", 0) > 2:
                            bt["is_occupied"] = True
                            bt["entry_time"] = now
                            bt["last_inside_motion"] = now
                            bt["alert_level"] = None
                            print(f"{GREEN}🚪 {loc} ENTER{RESET}")
                            bt["door_debounce"] = now
                            log_value = "ENTER"
                elif str(val).upper() in ["EXIT", "CLEAR"]:
                    if bt.get("is_occupied", False) and (now - last_debounce > DEBOUNCE_TIME):
                        if now - bt.get("entry_time", 0) > 2:
                            bt["is_occupied"] = False
                            bt["alert_level"] = None
                            bt["critical_sent"] = False
                            bt["humidity_alert_sent"] = False
                            print(f"{RED}🚪 {loc} EXIT{RESET}")
                            bt["door_debounce"] = now
                            log_value = "EXIT"

                if log_value:
                    log_to_excel("Proximity", loc, log_value, f"User {log_value}ED {loc}")
                room_states[loc] = bt

        # --- HUMIDITY ---
        elif s_type == "Humidity" and loc == "Bathroom":
            try:
                room_states["Bathroom"]["humidity"] = float(str(val).replace("%", ""))
            except:
                pass
            log_to_excel(s_type, loc, val, payload.get("status"))

        # --- TEMPERATURE ---
        elif s_type == "Temperature":
            location_temp = loc if loc else "Unknown"
            log_to_excel("Temperature", location_temp, val, payload.get("status"))

        # --- MMWAVE / HUMAN PRESENCE ---
        elif str(s_type).strip().lower() in ["mmwave", "human presence", "presence"]:
            val_str = str(val).strip().upper() if val is not None else ""
            hr = payload.get("heart_rate") or payload.get("hr")
            br = payload.get("breath_rate") or payload.get("br") or payload.get("respiration_rate")
            status_field = payload.get("status", "Active")

            # --- FALL DETECTION ---
            if val_str == "FALL_DETECTED":
                if loc not in room_states:
                    room_states[loc] = {"last_fall_alert": 0}
                last_alert_time = room_states[loc].get("last_fall_alert", 0)
                if now - last_alert_time < 60:
                    print(f"🛑 [SPAM FILTER] Ignoring duplicate Fall Alert for {loc}")
                    return
                fall_msg = f"🚨 EMERGENCY: Fall Detected at {loc} at {datetime.now().strftime('%H:%M:%S')}"
                send_email_alert("FALL DETECTION ALERT", fall_msg)
                room_states[loc]["last_fall_alert"] = now
                log_to_excel("ALERTS", loc, "CRITICAL", "FALL_DETECTED")
                print(f"{RED}{fall_msg}{RESET}")
                return

            # --- IN BED / OUT OF BED ---
            if val_str in ["IN BED", "1", "TRUE", "YES", "PRESENCE", "DETECTED"] or str(status_field).strip().upper() == "OCCUPIED":
                bed_state = "In Bed"
            elif val_str in ["OUT OF BED", "0", "FALSE", "NO", "NO_PRESENCE"] or str(status_field).strip().upper() == "EMPTY":
                bed_state = "Out of Bed"
            else:
                bed_state = None

            is_bedroom_stream = (loc and str(loc).lower() == "bedroom") or ("mmwave_bedroom" in msg.topic)
            if is_bedroom_stream or (hr is not None) or (br is not None):
                location_bedroom = "Bedroom"
                if bed_state is not None:
                    log_to_excel("mmWave(InBed)", location_bedroom, bed_state, status_field)
                if hr is not None:
                    try_val = int(hr) if str(hr).isdigit() else hr
                    log_to_excel("mmWave(HR)", location_bedroom, try_val, status_field)
                if br is not None:
                    try_val = int(br) if str(br).isdigit() else br
                    log_to_excel("mmWave(BR)", location_bedroom, try_val, status_field)
                print(f"📡 mmWave Bedroom -> HR={hr if hr else '-'} | BR={br if br else '-'} | State={bed_state if bed_state else (val if val else '-')}")

                # =========================================================
                # ✅ FIXED: BEDROOM VITALS ALERT RULING (OCCUPIED + 30s WAIT)
                # =========================================================
                bd = room_states["Bedroom"]

                # Parse HR/BR
                hr_val, br_val = None, None
                try:
                    if hr is not None and str(hr).strip() != "":
                        hr_val = float(hr)
                    if br is not None and str(br).strip() != "":
                        br_val = float(br)
                except:
                    pass

                hr_low = hr_val is not None and hr_val < BEDROOM_VITALS_HR_LOW
                br_low = br_val is not None and br_val < BEDROOM_VITALS_BR_LOW

                # RULE: only consider vitals if Bedroom is occupied AND in bed
                should_track_vitals = bd.get("is_occupied", False) and (bed_state == "In Bed")

                if should_track_vitals and (hr_low or br_low):
                    # Start pending timer on first detection
                    if not bd.get("low_vitals_pending", False):
                        bd["low_vitals_pending"] = True
                        bd["low_vitals_start"] = now
                        print(f"{YELLOW}⏳ Bedroom low vitals detected (pending {BEDROOM_VITALS_CONFIRM_SECONDS}s) ...{RESET}")
                    else:
                        # If low vitals stays low for 30s, then send (respect cooldown)
                        pending_for = now - bd.get("low_vitals_start", now)
                        if pending_for >= BEDROOM_VITALS_CONFIRM_SECONDS:
                            last_sent = bd.get("last_vitals_alert", 0)
                            if now - last_sent >= BEDROOM_VITALS_ALERT_COOLDOWN:
                                email_body = build_bedroom_vitals_email(hr_val, br_val, hr_low, br_low)
                                send_email_alert("BEDROOM VITAL SIGNS ALERT", email_body)

                                log_parts = []
                                if hr_low:
                                    log_parts.append(f"HR low ({hr_val} < {BEDROOM_VITALS_HR_LOW})")
                                if br_low:
                                    log_parts.append(f"BR low ({br_val} < {BEDROOM_VITALS_BR_LOW})")
                                log_to_excel("System", "Bedroom", "CRITICAL", " | ".join(log_parts))

                                bd["last_vitals_alert"] = now
                                bd["low_vitals_pending"] = False
                                bd["low_vitals_start"] = 0
                                print(f"{RED}🚨 BEDROOM VITALS ALERT SENT (after 30s confirm){RESET}")
                            else:
                                print("🛑 Bedroom vitals alert suppressed (cooldown active)")
                                # keep pending off to avoid sending immediately after cooldown
                                bd["low_vitals_pending"] = False
                                bd["low_vitals_start"] = 0
                else:
                    # Cancel pending if:
                    # - left bedroom (not occupied)
                    # - out of bed
                    # - vitals recovered
                    if bd.get("low_vitals_pending", False):
                        print(f"{CYAN}✅ Bedroom low vitals pending cancelled (left room / out of bed / recovered).{RESET}")
                    bd["low_vitals_pending"] = False
                    bd["low_vitals_start"] = 0

                room_states["Bedroom"] = bd
            else:
                presence = "PRESENCE" if bed_state == "In Bed" else "NO_PRESENCE" if bed_state == "Out of Bed" else str(val).upper() if val else "UNKNOWN"
                location_generic = loc if loc else "Unknown"
                log_to_excel("mmWave", location_generic, presence, status_field)
                print(f"📡 mmWave -> {location_generic}: {presence}")

    except Exception as e:
        print(f"⚠️ Error processing: {e}")

# ================= MAIN =================
client = mqtt.Client(mqtt.CallbackAPIVersion.VERSION2)
client.on_connect = on_connect
client.on_message = on_message

threading.Thread(target=bathroom_monitor_loop, daemon=True).start()
threading.Thread(target=report_monitor_loop, daemon=True).start()

print("✅ Central Logic Engine Running...")
resolved_broker = resolve_broker_address(BROKER_ADDRESS)
if resolved_broker != BROKER_ADDRESS:
    print(f"{CYAN}🔎 Broker resolved: {BROKER_ADDRESS} -> {resolved_broker}{RESET}")
else:
    print(f"{CYAN}🔎 Broker target: {BROKER_ADDRESS}{RESET}")

client.connect(resolved_broker, 1883, 60)
client.loop_forever()

